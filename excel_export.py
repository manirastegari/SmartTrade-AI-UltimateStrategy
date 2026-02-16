#!/usr/bin/env python3
"""
Excel Export Functionality for SmartTrade AI
Professional reporting and portfolio management
"""

import pandas as pd
import numpy as np
from datetime import datetime
import os
import subprocess
from openpyxl.utils import get_column_letter

def _clean_val(val, max_len=30000):
    """Aggressive sanitization for Excel export."""
    if pd.isna(val) or val is None:
        return ""
    
    # Preserve numbers
    if isinstance(val, (int, float)):
        return val
        
    val = str(val)
    
    # Prevent formula injection
    if val.startswith("="):
        val = "'" + val
        
    # Whitelist: Allow only standard printable characters (ASCII + common symbols)
    # This strips complex Unicode that Excel might hate (emojis, control chars, etc.)
    import re
    # Keep alphanumeric, punctuation, and common whitespace (space)
    # Strip everything else to be 100% safe
    val = re.sub(r'[^\x20-\x7E]', '', val) 
    
    if len(val) > max_len:
        val = val[:max_len] + "..."
        
    return val

def push_to_github(filename):
    """
    Automatically commit and push Excel results to GitHub
    
    Args:
        filename: The Excel file to commit
    
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        # Get the directory of the Excel file
        file_dir = os.path.dirname(os.path.abspath(filename)) if os.path.dirname(filename) else os.getcwd()
        file_name = os.path.basename(filename)
        
        # Git add the file
        subprocess.run(['git', 'add', file_name], cwd=file_dir, check=True, capture_output=True)
        
        # Git commit
        commit_message = f"Auto-export: {file_name} - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        subprocess.run(['git', 'commit', '-m', commit_message], cwd=file_dir, check=True, capture_output=True)
        
        # Git push
        subprocess.run(['git', 'push'], cwd=file_dir, check=True, capture_output=True)
        
        print(f"✅ Successfully pushed {file_name} to GitHub")
        return True
        
    except subprocess.CalledProcessError as e:
        # Git command failed - might be nothing to commit or network issue
        if b'nothing to commit' in e.stderr or b'nothing added to commit' in e.stderr:
            print(f"⚠️ No changes to commit for {filename}")
        else:
            print(f"⚠️ Git push failed: {e.stderr.decode() if e.stderr else str(e)}")
        return False
    except Exception as e:
        print(f"⚠️ Git push error: {str(e)}")
        return False

def export_analysis_to_excel(results, analysis_params=None, filename=None, auto_push_github=None, all_stocks_data=None, market_tradability=None, market_timing_signal=None, ai_top_picks=None, analysis_start_time=None, analysis_end_time=None, analysis_duration_minutes=None, ai_universe_context=None, day_assessment=None):
    """Export analysis results to Excel with multiple sheets
    
    Optimized for Premium Quality Universe (614 institutional-grade stocks)
    
    Args:
        results: Consensus/filtered recommendations list
        analysis_params: Analysis parameters string
        filename: Output filename (optional)
        auto_push_github: Auto-commit to GitHub. If None, falls back to SMARTTRADE_AUTO_PUSH env flag.
        all_stocks_data: Complete list of ALL analyzed stocks (NEW - for full dataset export)
        market_tradability: AI market tradability analysis (NEW - for AI insights)
        market_timing_signal: Market timing signal (NEW - BUY/WAIT/SELL signal)
        analysis_start_time: Analysis start timestamp
        analysis_end_time: Analysis end timestamp
        analysis_duration_minutes: Total analysis duration in minutes
        ai_universe_context: AI Phase 1 selection context (NEW)
        day_assessment: Market Day Advisor assessment with skip warnings (NEW - Phase 2)
    """
    
    if not results and not all_stocks_data:
        return None, "No results to export"
    
    # Generate filename if not provided
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"SmartTrade_Premium_Analysis_{timestamp}.xlsx"
    
    try:
        # Create Excel writer object
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            
            # Sheet 1: Summary Dashboard
            # Safe sheet creation wrapper
            def safe_create(func, *args, **kwargs):
                try:
                    func(*args, **kwargs)
                except Exception as e:
                    print(f"⚠️ Error creating sheet: {str(e)}")

            # Sheet 1: Summary Dashboard
            safe_create(create_summary_sheet,
                results,
                writer,
                analysis_params,
                all_stocks_count=len(all_stocks_data) if all_stocks_data else None,
                market_tradability=market_tradability,
                market_timing_signal=market_timing_signal,
                ai_top_picks=ai_top_picks,
                analysis_start_time=analysis_start_time,
                analysis_end_time=analysis_end_time,
                analysis_duration_minutes=analysis_duration_minutes,
                day_assessment=day_assessment,
                ai_universe_context=ai_universe_context
            )
            
            # Sheet 2: ALL ANALYZED STOCKS
            if all_stocks_data:
                safe_create(create_all_analyzed_sheet, all_stocks_data, writer)

            # Sheet 3: AI Top Picks
            if ai_top_picks:
                safe_create(create_ai_top_picks_sheet, ai_top_picks, writer)
            
            # Sheet 4: Ultimate Buy
            safe_create(create_recommendations_sheet, results, writer, 'ULTIMATE BUY', 'Ultimate_Buy')

            # Sheet 5: Strong Buy
            safe_create(create_recommendations_sheet, results, writer, 'STRONG BUY')
            
            # Sheet 6: All Buy Signals
            safe_create(create_recommendations_sheet, results, writer, ['ULTIMATE BUY', 'STRONG BUY', 'BUY', 'WEAK BUY'], 'All_Buy_Signals')
            
            # Sheet 7: Detailed Analysis
            safe_create(create_detailed_analysis_sheet, results, writer)
            
            # Sheet 8: Technical Indicators
            safe_create(create_technical_sheet, results, writer)
            
            # Sheet 9: Risk Analysis
            safe_create(create_risk_analysis_sheet, results, writer)
            
            # Sheet 10: Sector Analysis
            safe_create(create_sector_analysis_sheet, results, writer)
            
            # Sheet 11: Performance Metrics
            safe_create(create_performance_sheet, results, writer)
        
        # Auto-push to GitHub if requested (ENABLED BY DEFAULT)
        env_push = os.getenv('SMARTTRADE_AUTO_PUSH', 'true').lower() in ('1', 'true', 'yes')
        should_push = auto_push_github if auto_push_github is not None else env_push
        if should_push:
            print("📤 Auto-pushing Excel report to GitHub...")
            push_to_github(filename)
        
        total_analyzed = len(all_stocks_data) if all_stocks_data else len(results)
        consensus_count = len(results) if results else 0
        return filename, f"Successfully exported {total_analyzed} analyzed stocks ({consensus_count} consensus picks) to {filename}"
        
    except Exception as e:
        return None, f"Export failed: {str(e)}"

def create_summary_sheet(results, writer, analysis_params, all_stocks_count=None, market_tradability=None, market_timing_signal=None, ai_top_picks=None, analysis_start_time=None, analysis_end_time=None, analysis_duration_minutes=None, ai_universe_context=None, day_assessment=None):
    """Create summary dashboard sheet
    
    Args:
        results: Consensus picks list
        writer: Excel writer
        analysis_params: Analysis parameters string
        all_stocks_count: Total number of stocks analyzed (NEW - to show complete picture)
        market_tradability: AI market tradability analysis (NEW - for AI insights)
        market_timing_signal: Market timing signal (NEW - BUY/WAIT/SELL signal)
        analysis_start_time: Analysis start timestamp
        analysis_end_time: Analysis end timestamp
        analysis_duration_minutes: Total analysis duration in minutes
        ai_universe_context: AI Phase 1 selection context (NEW)
        day_assessment: Market Day Advisor assessment with skip warnings (NEW - Phase 2)
    """
    
    # Check if this is consensus format
    is_consensus = isinstance(results, list) and results and 'strategies_agreeing' in results[0]
    
    if is_consensus:
        # New consensus format
        total_stocks = len(results)
        strong_buy = len([r for r in results if r.get('recommendation') == 'STRONG BUY'])
        ultimate_buy = len([r for r in results if r.get('recommendation') == 'ULTIMATE BUY'])
        buy = len([r for r in results if r.get('recommendation') == 'BUY'])
        weak_buy = len([r for r in results if r.get('recommendation') == 'WEAK BUY'])
        hold = len([r for r in results if r.get('recommendation') == 'HOLD'])
        sell_signals = 0  # Not used in consensus
        
        # Consensus-specific stats
        tier_5 = len([r for r in results if r.get('strategies_agreeing') == 5])
        tier_4 = len([r for r in results if r.get('strategies_agreeing') == 4])
        tier_3 = len([r for r in results if r.get('strategies_agreeing') == 3])
        tier_2 = len([r for r in results if r.get('strategies_agreeing') == 2])
        
        avg_quality = np.mean([r.get('quality_score', 0) for r in results]) if results else 0
        top_performer = max(results, key=lambda x: x.get('quality_score', 0)).get('symbol', 'N/A') if results else 'N/A'
        
        # CRITICAL FIX: Show both total analyzed and consensus picks
        total_analyzed_display = all_stocks_count if all_stocks_count else total_stocks
        consensus_picks_display = total_stocks
        
        # Create consensus summary
        summary_data = {
            'Metric': [
                '🚦 TRADING DAY SUMMARY',
                'Trading Action',
                'Skip Today?',
                'Confidence',
                'Honest Assessment',
                'Position Sizing Recommendation',
                'Strategy Focus',
                'Next Check',
                '━━━━━━━━━━━━━━━━━━━━━━━━━━',
                'Analysis Start Time',
                'Analysis End Time',
                'Total Duration (minutes)',
                'Analysis Type',
                'Universe Type',
                'Total Stocks Analyzed',
                'Consensus Picks (2+ Agreement)',
                '━━━━━━━━━━━━━━━━━━━━━━━━━━',
                '📊 MARKET TIMING SIGNAL',
                'ACTION',
                'Signal',
                'Position Sizing',
                'Confidence',
                'VIX Level',
                'SPY 1D Return',
                'Market Regime',
                'Reason',
                '━━━━━━━━━━━━━━━━━━━━━━━━━━',
                '🧠 AI REASONING TRACE',
                'Reasoning Logic',
                '',
                '🤖 AI MARKET ANALYSIS',
                'AI Trade Recommendation',
                'AI Confidence Level',
                'AI Market Summary',
                '━━━━━━━━━━━━━━━━━━━━━━━━━━',
                '🌍 AI UNIVERSE SELECTION',
                'Selection Strategy',
                'Focus Sectors',
                '',
                '',
                '🎯 AI TOP PICKS SUMMARY',
                'AI Picks Generated',
                'AI Key Insight',
                'Top Ranked Symbols',
                '━━━━━━━━━━━━━━━━━━━━━━━━━━',
                '5/5 Agreement (ULTIMATE BUY)',
                '4/5 Agreement (STRONG BUY)',
                '3/5 Agreement (BUY)',
                '2/5 Agreement (WEAK BUY)',
                'Average Quality Score',
                'Top Performer',
                '━━━━━━━━━━━━━━━━━━━━━━━━━━',
                'Methodology',
                'Risk Management',
                'Analysis Parameters'
            ],
            'Value': [
                '',  # Section header
                _clean_val(f"{'🔴 SKIP' if day_assessment and day_assessment.get('skip_today') else '🟡 CAUTION' if day_assessment and day_assessment.get('warning_level') == 'YELLOW' else '🟢 TRADE'}" if day_assessment else 'N/A'),
                _clean_val('YES - Wait for better conditions' if day_assessment and day_assessment.get('skip_today') else 'NO - Conditions acceptable'),
                _clean_val(f"{day_assessment.get('confidence', 0):.0f}%" if day_assessment else 'N/A'),
                _clean_val(day_assessment.get('honest_assessment', 'Day assessment not available') if day_assessment else 'Day assessment not available'),
                _clean_val(day_assessment.get('position_sizing', 'N/A') if day_assessment else 'N/A'),
                _clean_val(day_assessment.get('strategy_focus', 'N/A') if day_assessment else 'N/A'),
                _clean_val(day_assessment.get('next_check', 'N/A') if day_assessment else 'N/A'),
                '',  # Separator
                _clean_val(analysis_start_time if analysis_start_time else datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                _clean_val(analysis_end_time if analysis_end_time else datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                _clean_val(f"{analysis_duration_minutes} minutes" if analysis_duration_minutes else "N/A"),
                'Premium Ultimate Strategy - 5-Perspective Consensus',
                _clean_val(f"Interactive AI Selection ({ai_universe_context.get('reasoning')[:50]}...)" if ai_universe_context else 'Premium Quality Universe (614 institutional-grade stocks)'),
                _clean_val(f"{total_analyzed_display} stocks (AI Focused)" if ai_universe_context else f"{total_analyzed_display} stocks (complete analysis)"),
                _clean_val(f"{consensus_picks_display} stocks (filtered by multi-strategy agreement)"),
                '',  # Separator
                '',  # Section header
                _clean_val(market_timing_signal.get('action', 'N/A') if market_timing_signal else 'N/A'),
                _clean_val(market_timing_signal.get('signal', 'N/A') if market_timing_signal else 'N/A'),
                _clean_val(market_timing_signal.get('position_sizing', 'N/A') if market_timing_signal else 'N/A'),
                _clean_val(f"{market_timing_signal.get('confidence', 0)}%" if market_timing_signal else 'N/A'),
                _clean_val((f"{market_timing_signal.get('vix_level')}" if market_timing_signal and market_timing_signal.get('vix_level') is not None else 'N/A')),
                _clean_val((f"{market_timing_signal.get('spy_return_1d')}%" if market_timing_signal and market_timing_signal.get('spy_return_1d') is not None else 'N/A')),
                _clean_val(market_timing_signal.get('market_regime', 'N/A') if market_timing_signal else 'N/A'),
                _clean_val(market_timing_signal.get('brief_reason', 'No timing signal available') if market_timing_signal else 'No timing signal available'),
                '',  # Separator
                '',  # Section header
                '🧠 AI REASONING TRACE',
                _clean_val(ai_top_picks.get('reasoning_trace', 'N/A') if ai_top_picks else 'N/A'),
                '',
                '🤖 AI MARKET ANALYSIS',
                _clean_val(market_tradability.get('trade_recommendation', 'N/A') if market_tradability else 'N/A'),
                _clean_val(f"{market_tradability.get('confidence', 0):.0f}%" if market_tradability else 'N/A'),
                _clean_val(market_tradability.get('brief_summary', 'AI analysis not available') if market_tradability else 'AI analysis not available'),
                '',
                '🌍 AI UNIVERSE SELECTION',
                _clean_val(f"Strategy: {ai_universe_context.get('reasoning')}" if ai_universe_context else 'N/A'),
                _clean_val(f"Focus Sectors: {', '.join(ai_universe_context.get('focus_sectors', []))}" if ai_universe_context else 'N/A'),
                '',
                '',
                _clean_val(f"{ai_top_picks.get('total_recommended', 0)} of {ai_top_picks.get('total_analyzed', 0)} candidates" if ai_top_picks else 'AI selection unavailable'),
                _clean_val(ai_top_picks.get('key_insight', 'N/A') if ai_top_picks else 'N/A'),
                _clean_val(', '.join([p.get('symbol', 'N/A') for p in ai_top_picks.get('ai_top_picks', [])[:5]]) if ai_top_picks else 'N/A'),
                '',  # Separator
                _clean_val(f"{tier_5} stocks (all 5 perspectives agree)"),
                _clean_val(f"{tier_4} stocks (4 of 5 agree)"),
                _clean_val(f"{tier_3} stocks (3 of 5 agree)"),
                _clean_val(f"{tier_2} stocks (2 of 5 agree)"),
                _clean_val(f"{avg_quality:.1f}/100"),
                _clean_val(f"{top_performer} ({max([r.get('quality_score', 0) for r in results]):.0f}/100)" if results else 'N/A'),
                '',  # Separator
                '15 Quality Metrics: Fundamentals 40%, Momentum 30%, Risk 20%, Sentiment 10%',
                'Guardrails: DISABLED (pre-screened) | Regime Filters: RELAXED',
                _clean_val(str(analysis_params) if analysis_params else 'Premium Ultimate Strategy - 5-Perspective Consensus')
            ]
        }
    else:
        # Old format
        total_stocks = len(results)
        strong_buy = len([r for r in results if r.get('recommendation') == 'STRONG BUY'])
        buy = len([r for r in results if r.get('recommendation') == 'BUY'])
        weak_buy = len([r for r in results if r.get('recommendation') == 'WEAK BUY'])
        hold = len([r for r in results if r.get('recommendation') == 'HOLD'])
        sell_signals = len([r for r in results if r.get('recommendation', '').endswith('SELL')])
        
        summary_data = {
            'Metric': [
                'Analysis Date',
                'Universe Type',
                'Total Stocks Analyzed',
                'Strong Buy Recommendations',
                'Buy Recommendations', 
                'Weak Buy Recommendations',
                'Hold Recommendations',
                'Sell Signals',
                'Success Rate (%)',
                'Average Score',
                'Top Performer',
                'Risk Management',
                'Analysis Parameters'
            ],
            'Value': [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                'Premium Quality Universe (614 institutional-grade stocks)',
                total_stocks,
                strong_buy,
                buy,
                weak_buy,
                hold,
                sell_signals,
                f"{(strong_buy + buy + weak_buy) / total_stocks * 100:.1f}%" if total_stocks > 0 else "0%",
                f"{np.mean([r.get('overall_score', 0) for r in results]):.1f}",
                max(results, key=lambda x: x.get('overall_score', 0)).get('symbol', 'N/A') if results else 'N/A',
                'Guardrails: DISABLED (pre-screened) | Regime Filters: RELAXED',
                str(analysis_params) if analysis_params else 'Ultimate Strategy 4-Perspective Consensus'
            ]
        }
    
    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, sheet_name='Summary', index=False)

def create_all_analyzed_sheet(all_stocks_data, writer):
    """
    Create sheet showing ALL analyzed stocks (not just consensus picks)
    
    CRITICAL: This shows the complete 613-stock analysis
    Now includes Enhanced Signals (VWAP, Sector, S/R, RSI(2), ATR) for 20%+ accuracy
    
    Args:
        all_stocks_data: List of dict with all analyzed stocks and their metrics
        writer: Excel writer object
    """
    if not all_stocks_data:
        return
    
    # Prepare data for Excel
    excel_data = []
    
    for stock in all_stocks_data:
        # ML predictions (if available)
        ml_prob = stock.get('ml_probability')
        ml_return = stock.get('ml_expected_return')
        ml_conf = stock.get('ml_confidence')
        
        # Enhanced signals (NEW - 20%+ accuracy improvement)
        enhanced = stock.get('enhanced', {})
        
        excel_data.append({
            'Symbol': stock.get('symbol', 'N/A'),
            'Sector': stock.get('sector', 'Unknown'),
            'Quality Score': stock.get('quality_score', 0),
            'Current Price': stock.get('current_price', 0),
            
            # ML Predictions (NEW - Critical for users to see)
            'ML Probability %': round(ml_prob * 100, 1) if ml_prob is not None else None,
            'ML Expected Return %': round(ml_return, 1) if ml_return is not None else None,
            'ML Confidence %': round(ml_conf * 100, 1) if ml_conf is not None else None,
            
            # ENHANCED SIGNALS (NEW - 20%+ Accuracy Boost)
            'Enhancement Score': _clean_val(enhanced.get('enhancement_score')),
            'Enhancement Signal': _clean_val(enhanced.get('enhancement_signal', 'N/A')),
            'Confirmations': _clean_val(enhanced.get('confirmation_count', 0)),
            
            # VWAP Analysis
            'VWAP': _clean_val(enhanced.get('vwap')),
            'VWAP Signal': _clean_val(enhanced.get('vwap_signal', 'N/A')),
            'Breakout Confirmed': _clean_val('YES' if enhanced.get('breakout_confirmed') else 'NO'),
            
            # Entry Zone (Support/Resistance)
            'Entry Zone': _clean_val(enhanced.get('entry_zone', 'N/A')),
            'Entry Score': _clean_val(enhanced.get('entry_score')),
            'Entry Timing': _clean_val(enhanced.get('entry_timing', 'N/A')),
            'Nearest Support': _clean_val(enhanced.get('nearest_support')),
            'Nearest Resistance': _clean_val(enhanced.get('nearest_resistance')),
            'Risk/Reward': _clean_val(enhanced.get('risk_reward_ratio')),
            
            # Mean Reversion (RSI-2)
            'RSI(2)': _clean_val(enhanced.get('rsi_2')),
            'Reversion Signal': _clean_val(enhanced.get('reversion_signal', 'N/A')),
            'Bounce Probability %': _clean_val(enhanced.get('bounce_probability')),
            'Is Bounce Setup': _clean_val('YES' if enhanced.get('is_bounce_setup') else 'NO'),
            
            # ATR Stop Loss
            'Recommended Stop': _clean_val(enhanced.get('recommended_stop')),
            'Stop Loss %': _clean_val(enhanced.get('stop_loss_pct')),
            'Target (2:1 R:R)': _clean_val(enhanced.get('target_2r')),
            'Volatility Regime': _clean_val(enhanced.get('volatility_regime', 'N/A')),
            'Position Size': _clean_val(enhanced.get('position_size_suggestion', 'N/A')),
            
            # Sector Rotation
            'Sector Rank': _clean_val(enhanced.get('sector_rank')),
            'Sector Tier': _clean_val(enhanced.get('sector_tier', 'N/A')),
            
            # Fundamentals
            'P/E Ratio': _clean_val(stock.get('pe_ratio')),
            'Revenue Growth %': _clean_val(stock.get('revenue_growth')),
            'Profit Margin %': _clean_val(stock.get('profit_margin')),
            'ROE %': _clean_val(stock.get('roe')),
            'Debt/Equity': _clean_val(stock.get('debt_equity')),
            'Fund Score': _clean_val(stock.get('fundamentals_score')),
            'Fund Grade': _clean_val(stock.get('fundamentals_grade', 'N/A')),
            
            # Momentum
            'RSI': _clean_val(stock.get('rsi_14')),
            'Price Trend': _clean_val(stock.get('price_trend', 'N/A')),
            'Volume Trend': _clean_val(stock.get('volume_trend', 'N/A')),
            'Momentum Score': _clean_val(stock.get('momentum_score')),
            'Momentum Grade': _clean_val(stock.get('momentum_grade', 'N/A')),
            'MA 50': _clean_val(stock.get('ma_50')),
            'MA 200': _clean_val(stock.get('ma_200')),
            'Volume Ratio': _clean_val(stock.get('volume_ratio')),
            
            # Risk
            'Beta': _clean_val(stock.get('beta')),
            'Volatility %': _clean_val(stock.get('volatility')),
            'Sharpe Ratio': _clean_val(stock.get('sharpe_ratio')),
            'Max Drawdown %': _clean_val(stock.get('max_drawdown')),
            'VaR 95%': _clean_val(stock.get('var_95')),
            'Risk Score': _clean_val(stock.get('risk_score')),
            'Risk Grade': _clean_val(stock.get('risk_grade', 'N/A')),
            'Risk Level': _clean_val(stock.get('risk_level', 'N/A')),
            
            # Technical
            'MACD': _clean_val(stock.get('macd')),
            'MACD Signal': _clean_val(stock.get('macd_signal')),
            'MACD Histogram': _clean_val(stock.get('macd_hist')),
            'Bollinger Position %': _clean_val(stock.get('bollinger_position')),
            'Bollinger Upper': _clean_val(stock.get('bollinger_upper')),
            'Bollinger Lower': _clean_val(stock.get('bollinger_lower')),
            'Support Level': _clean_val(stock.get('support_level')),
            'Resistance Level': _clean_val(stock.get('resistance_level')),
            'Volume SMA': _clean_val(stock.get('volume_sma')),
            'Technical Score': _clean_val(stock.get('technical_score')),
            
            # Sentiment
            'Sentiment Score': _clean_val(stock.get('sentiment_score')),
            'Sentiment Grade': _clean_val(stock.get('sentiment_grade', 'N/A')),
            'Target Upside %': _clean_val(stock.get('target_upside')),
            'Institutional Ownership %': _clean_val(stock.get('institutional_ownership')),
            'Analyst Rating': _clean_val(stock.get('analyst_rating', 'N/A'))
        })
    
    # Create DataFrame
    df = pd.DataFrame(excel_data)
    
    # Sort by quality score (highest first)
    df = df.sort_values('Quality Score', ascending=False)
    
    # Write to Excel
    df.to_excel(writer, sheet_name='All_Analyzed_Stocks', index=False)
    
    # Get the worksheet
    worksheet = writer.sheets['All_Analyzed_Stocks']
    
    # Auto-adjust column widths
    for idx, col in enumerate(df.columns, start=1):
        max_len = max(df[col].astype(str).apply(len).max(), len(col)) + 2
        column_letter = get_column_letter(idx)
        worksheet.column_dimensions[column_letter].width = min(max_len, 30)
    
    # Add conditional formatting for Quality Score
    from openpyxl.styles import PatternFill
    from openpyxl.formatting.rule import CellIsRule
    
    # Green for high scores (80+)
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    worksheet.conditional_formatting.add(f'B2:B{len(df)+1}', 
        CellIsRule(operator='greaterThan', formula=['80'], fill=green_fill))
    
    # Yellow for medium scores (60-79)
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    worksheet.conditional_formatting.add(f'B2:B{len(df)+1}', 
        CellIsRule(operator='between', formula=['60', '79'], fill=yellow_fill))
    
    # Red for low scores (<60)
    red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
    worksheet.conditional_formatting.add(f'B2:B{len(df)+1}', 
        CellIsRule(operator='lessThan', formula=['60'], fill=red_fill))

def create_ai_top_picks_sheet(ai_top_picks, writer):
    """Create dedicated sheet highlighting AI-selected top opportunities."""
    if not ai_top_picks:
        return

    picks = ai_top_picks.get('ai_top_picks', [])

    # Summary block always written so the sheet exists even if no picks.
    summary_df = pd.DataFrame({
        'Metric': [
            'Brief Summary',
            'Key Insight',
            'Total Candidates Evaluated',
            'Total Recommended'
        ],
        'Value': [
            _clean_val(ai_top_picks.get('brief_summary', 'No AI summary available')),
            _clean_val(ai_top_picks.get('key_insight', 'N/A')),
            _clean_val(ai_top_picks.get('total_analyzed', 0)),
            _clean_val(ai_top_picks.get('total_recommended', 0))
        ]
    })
    summary_df.to_excel(writer, sheet_name='AI_Top_Picks', index=False)

    if picks:
        # Prepare data for DataFrame
        data = []
        for pick in picks: # Changed picks_list to picks
            current_price = pick.get('current_price', 0)
            # Calculate Buy Zone and Take Profit if not present
            if current_price and current_price > 0:
                buy_zone = pick.get('buy_zone', f"${current_price:.2f} - ${current_price*1.02:.2f}")
                take_profit = pick.get('take_profit', f"${current_price*1.15:.2f} (+15%)")
                price_display = f"${current_price:.2f}"
            else:
                buy_zone = pick.get('buy_zone', 'N/A')
                take_profit = pick.get('take_profit', 'N/A')
                price_display = 'N/A'
            
            # Robust confidence parsing: handle int, float (0-1 or 0-100), or string
            raw_conf = pick.get('confidence', 0)
            try:
                conf_val = float(str(raw_conf).replace('%', ''))
                if conf_val <= 1.0 and conf_val > 0:
                    conf_val = int(conf_val * 100)  # Convert 0-1 scale to percentage
                else:
                    conf_val = int(conf_val)
            except (ValueError, TypeError):
                conf_val = 0
            
            data.append({
                'Rank': _clean_val(pick.get('rank', 0)),
                'Symbol': _clean_val(pick.get('symbol', 'N/A')),
                'Current Price': _clean_val(price_display),
                'Buy Zone': _clean_val(buy_zone),
                'Take Profit': _clean_val(take_profit),
                'AI Confidence': _clean_val(f"{conf_val}%"),
                'Macro Fit': _clean_val(pick.get('macro_fit', 'N/A')),
                'Reasoning': _clean_val(pick.get('reasoning', '') or pick.get('why_selected', ''))
            })
            
        df = pd.DataFrame(data)
        
        # Write to Excel
        start_row = len(summary_df) + 2 # Added start_row logic
        df.to_excel(writer, sheet_name='AI_Top_Picks', index=False, startrow=start_row) # Changed sheet name to match summary

    worksheet = writer.sheets['AI_Top_Picks'] # Ensure worksheet is defined even if picks is empty

    # Adjust column widths
    column_widths = {
        'A': 6,   # Rank
        'B': 10,  # Symbol
        'B': 10,  # Symbol
        'C': 12,  # Current Price
        'D': 20,  # Buy Zone
        'E': 20,  # Take Profit
        'F': 15,  # AI Confidence
        'G': 15,  # Macro Fit
        'H': 60   # Reasoning
    }
    for col_letter, width in column_widths.items():
        worksheet.column_dimensions[col_letter].width = width
    
    # Auto-fit remaining columns if any, or adjust existing ones if content is longer
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            try:
                cell_length = len(str(cell.value)) if cell.value is not None else 0
            except Exception:
                cell_length = 0
            max_length = max(max_length, cell_length)
        # Only apply if auto-fit length is greater than manually set width, or if column wasn't manually set
        if column_letter not in column_widths or max_length + 2 > column_widths[column_letter]:
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 60)

def create_recommendations_sheet(results, writer, recommendation_types, sheet_name=None):
    """Create recommendations sheet (supports both old and new consensus format)"""
    
    if isinstance(recommendation_types, str):
        recommendation_types = [recommendation_types]
        sheet_name = sheet_name or recommendation_types[0].replace(' ', '_')
    else:
        sheet_name = sheet_name or 'Recommendations'
    
    # Filter results
    filtered_results = [r for r in results if r.get('recommendation') in recommendation_types]
    
    if not filtered_results:
        # Skip creating empty sheet
        return
    
    # Create recommendations data (handle both old and new consensus formats)
    recommendations_data = []
    for result in filtered_results:
        # Check if this is consensus format (has strategies_agreeing)
        is_consensus = 'strategies_agreeing' in result
        
        if is_consensus:
            # New Premium Ultimate Strategy format
            fund = result.get('fundamentals', {})
            mom = result.get('momentum', {})
            risk = result.get('risk', {})
            sent = result.get('sentiment', {})
            
            # ML predictions
            ml_prob = result.get('ml_probability')
            ml_return = result.get('ml_expected_return')
            ml_conf = result.get('ml_confidence')
            ultimate_score = result.get('ultimate_score')
            
            recommendations_data.append({
                'Symbol': _clean_val(result.get('symbol', '')),
                'Recommendation': _clean_val(result.get('recommendation', '')),
                'Agreement': _clean_val(f"{result.get('strategies_agreeing', 0)}/5"),
                
                # Ultimate Score (NEW - combines all layers)
                'Ultimate Score': _clean_val(ultimate_score if ultimate_score is not None else 'N/A'),
                # Entry Score (NEW - regime-aware entry suitability)
                'Entry Score': _clean_val(result.get('entry_score', 'N/A')),
                
                'Quality Score': _clean_val(result.get('quality_score', 0)),
                'Consensus Score': _clean_val(result.get('consensus_score', 0)),
                'Confidence': _clean_val(f"{result.get('confidence', 0) * 100:.0f}%"),
                
                # ML Predictions (NEW)
                'ML Probability': _clean_val(f"{ml_prob * 100:.1f}%" if ml_prob is not None else 'N/A'),
                'ML Expected Return': _clean_val(f"{ml_return:+.1f}%" if ml_return is not None else 'N/A'),
                'ML Confidence': _clean_val(f"{ml_conf * 100:.1f}%" if ml_conf is not None else 'N/A'),
                
                # AI Validation (NEW)
                'AI Validation': _clean_val(result.get('ai_validation', 'N/A')),
                'AI Risk Level': _clean_val(result.get('ai_risk_level', 'N/A')),
                'AI Profit Potential': _clean_val(result.get('ai_profit_potential', 'N/A')),
                'News Sentiment': _clean_val(result.get('ai_news_sentiment', 'N/A')),
                'AI Verdict': _clean_val(result.get('ai_verdict', 'N/A')),
                
                # Catalyst Analysis (NEW)
                'Catalyst Score': _clean_val(result.get('catalyst_score', 'N/A')),
                'Earnings Outlook': _clean_val(result.get('earnings_outlook', 'N/A')),
                'Top Catalysts': _clean_val(' | '.join(result.get('growth_catalysts', [])[:3]) if result.get('growth_catalysts') else 'N/A'),
                
                'Current Price': _clean_val(f"${result.get('current_price', 0):.2f}"),
                'Fundamentals': _clean_val(f"{fund.get('grade', 'N/A')} ({fund.get('score', 0):.0f})"),
                'Momentum': _clean_val(f"{mom.get('grade', 'N/A')} ({mom.get('score', 0):.0f})"),
                'Risk': _clean_val(f"{risk.get('grade', 'N/A')} ({risk.get('score', 0):.0f})"),
                'Sentiment': _clean_val(f"{sent.get('grade', 'N/A')} ({sent.get('score', 0):.0f})"),
                'Perspectives': _clean_val(', '.join(result.get('agreeing_perspectives', []))),
                'P/E Ratio': _clean_val(fund.get('pe_ratio', 'N/A')),
                'Revenue Growth': _clean_val(f"{fund.get('revenue_growth', 0)*100:.1f}%" if fund.get('revenue_growth') else 'N/A'),
                'Beta': _clean_val(risk.get('beta', 'N/A'))
            })
        else:
            # Old format
            recommendations_data.append({
                'Symbol': result.get('symbol', ''),
                'Company': result.get('company_name', ''),
                'Recommendation': result.get('recommendation', ''),
                'Current Price': f"${result.get('current_price', 0):.2f}",
                'Predicted Return': f"{result.get('prediction', 0) * 100:.1f}%",
                'Overall Score': result.get('overall_score', 0),
                'Technical Score': result.get('technical_score', 0),
                'Fundamental Score': result.get('fundamental_score', 0),
                'Risk Level': result.get('risk_level', ''),
                'Sector': result.get('sector', ''),
                'Market Cap': result.get('market_cap', 0),
                'Volume': result.get('volume', 0),
                'P/E Ratio': result.get('pe_ratio', 0),
                'Confidence': f"{result.get('confidence', 0) * 100:.1f}%"
            })
    
    recommendations_df = pd.DataFrame(recommendations_data)
    
    # Sort by appropriate column
    if recommendations_data and 'Quality Score' in recommendations_data[0]:
        recommendations_df = recommendations_df.sort_values('Quality Score', ascending=False)
    elif 'Overall Score' in recommendations_df.columns:
        recommendations_df = recommendations_df.sort_values('Overall Score', ascending=False)
    
    recommendations_df.to_excel(writer, sheet_name=sheet_name, index=False)

def create_detailed_analysis_sheet(results, writer):
    """Create detailed analysis sheet with all metrics"""
    if not results:
        empty_df = pd.DataFrame({'Message': ['No analysis results available']})
        empty_df.to_excel(writer, sheet_name='Detailed_Analysis', index=False)
        return

    detailed_data = []
    is_consensus = 'fundamentals' in results[0]

    for result in results:
        if is_consensus:
            fund = result.get('fundamentals', {})
            mom = result.get('momentum', {})
            risk = result.get('risk', {})
            tech = result.get('technical', {})
            sent = result.get('sentiment', {})
            
            # ML predictions
            ml_prob = result.get('ml_probability')
            ml_return = result.get('ml_expected_return')
            ml_conf = result.get('ml_confidence')
            ml_features = result.get('ml_feature_importance', {})
            ultimate_score = result.get('ultimate_score')
            
            # Get top ML feature
            top_ml_feature = 'N/A'
            if ml_features:
                top = max(ml_features.items(), key=lambda x: abs(x[1]))
                top_ml_feature = f"{top[0]}: {top[1]:.2f}"

            detailed_data.append({
                'Symbol': _clean_val(result.get('symbol', '')),
                'Recommendation': _clean_val(result.get('recommendation', '')),
                'Agreement': _clean_val(f"{result.get('strategies_agreeing', 0)}/5"),
                
                # Ultimate Score (NEW - combines Quality + Consensus + ML)
                'Ultimate Score': _clean_val(ultimate_score if ultimate_score is not None else None),
                
                'Consensus Score': _clean_val(result.get('consensus_score', result.get('overall_score', 0))),
                'Quality Score': _clean_val(result.get('quality_score', 0)),
                'Confidence %': _clean_val(result.get('confidence', 0) * 100),
                
                # ML Predictions (NEW)
                'ML Probability %': _clean_val(round(ml_prob * 100, 1) if ml_prob is not None else None),
                'ML Expected Return %': _clean_val(round(ml_return, 1) if ml_return is not None else None),
                'ML Confidence %': _clean_val(round(ml_conf * 100, 1) if ml_conf is not None else None),
                'ML Top Driver': _clean_val(top_ml_feature),
                
                # AI Validation (NEW)
                'AI Validation': _clean_val(result.get('ai_validation', 'N/A')),
                'AI Risk Level': _clean_val(result.get('ai_risk_level', 'N/A')),
                'AI Profit Potential': _clean_val(result.get('ai_profit_potential', 'N/A')),
                'News Sentiment': _clean_val(result.get('ai_news_sentiment', 'N/A')),
                'AI Hidden Risks': _clean_val(result.get('ai_hidden_risks', 'N/A')),
                'AI Verdict': _clean_val(result.get('ai_verdict', 'N/A')),
                
                # Catalyst Analysis (NEW)
                'Catalyst Score': _clean_val(result.get('catalyst_score', 'N/A')),
                'Earnings Outlook': _clean_val(result.get('earnings_outlook', 'N/A')),
                'Growth Catalysts': _clean_val(' | '.join(result.get('growth_catalysts', [])[:3]) if result.get('growth_catalysts') else 'N/A'),
                'Catalyst Risks': _clean_val(' | '.join(result.get('catalyst_risks', [])[:3]) if result.get('catalyst_risks') else 'N/A'),
                'Sentiment Summary': _clean_val(result.get('sentiment_summary', 'N/A')),
                'Catalyst Summary': _clean_val(result.get('catalyst_summary', 'N/A')),
                
                'Sector': _clean_val(result.get('sector', 'Unknown')),
                'Current Price': _clean_val(result.get('current_price', 0)),
                'Fundamentals Score': _clean_val(fund.get('score', 0)),
                'Fundamentals Grade': _clean_val(fund.get('grade', 'N/A')),
                'P/E Ratio': _clean_val(fund.get('pe_ratio')),
                'Revenue Growth %': _clean_val(fund.get('revenue_growth')),
                'Profit Margin %': _clean_val(fund.get('profit_margin')),
                'ROE %': _clean_val(fund.get('roe')),
                'Debt/Equity': _clean_val(fund.get('debt_equity')),
                'Momentum Score': _clean_val(mom.get('score', 0)),
                'Momentum Grade': _clean_val(mom.get('grade', 'N/A')),
                'RSI': _clean_val(mom.get('rsi')),
                'Price Trend': _clean_val(mom.get('price_trend')),
                'Volume Trend': _clean_val(mom.get('volume_trend')),
                'Relative Strength %': _clean_val(mom.get('relative_strength')),
                'MA 50': _clean_val(mom.get('ma_50')),
                'MA 200': _clean_val(mom.get('ma_200')),
                'Technical Score': _clean_val(tech.get('score', 0)),
                'Technical Grade': _clean_val(tech.get('grade', 'N/A')),
                'MACD': _clean_val(tech.get('macd')),
                'MACD Signal': _clean_val(tech.get('macd_signal')),
                'MACD Histogram': _clean_val(tech.get('macd_hist')),
                'Bollinger Position %': _clean_val(tech.get('bollinger_position')),
                'Support Level': _clean_val(tech.get('support')),
                'Resistance Level': _clean_val(tech.get('resistance')),
                'Risk Score': _clean_val(risk.get('score', 0)),
                'Risk Grade': _clean_val(risk.get('grade', 'N/A')),
                'Risk Level': _clean_val(risk.get('risk_level', '')),
                'Volatility %': _clean_val(risk.get('volatility')),
                'Max Drawdown %': _clean_val(risk.get('max_drawdown')),
                'VaR 95% %': _clean_val(risk.get('var_95')),
                'Sharpe Ratio': _clean_val(risk.get('sharpe_ratio')),
                'Sentiment Score': _clean_val(sent.get('score', 0)),
                'Sentiment Grade': _clean_val(sent.get('grade', 'N/A')),
                'Target Upside %': _clean_val(sent.get('target_upside')),
                'Institutional Ownership %': _clean_val(sent.get('institutional_ownership')),
                'Analyst Rating': _clean_val(sent.get('analyst_rating', 'N/A'))
            })
        else:
            detailed_data.append({
                'Symbol': result.get('symbol', ''),
                'Company': result.get('company_name', ''),
                'Recommendation': result.get('recommendation', ''),
                'Current Price': result.get('current_price', 0),
                'Predicted Return %': result.get('prediction', 0) * 100,
                'Confidence %': result.get('confidence', 0) * 100,
                'Overall Score': result.get('overall_score', 0),
                'Technical Score': result.get('technical_score', 0),
                'Fundamental Score': result.get('fundamental_score', 0),
                'Sentiment Score': result.get('sentiment_score', 0),
                'Risk Level': result.get('risk_level', ''),
                'Sector': result.get('sector', ''),
                'Market Cap': result.get('market_cap', 0),
                'Volume': result.get('volume', 0),
                'P/E Ratio': result.get('pe_ratio', 0),
                'P/B Ratio': result.get('pb_ratio', 0),
                'ROE %': result.get('roe', 0),
                'Debt/Equity': result.get('debt_equity', 0),
                'Revenue Growth %': result.get('revenue_growth', 0),
                'Earnings Growth %': result.get('earnings_growth', 0),
                'RSI': result.get('rsi', 0),
                'MACD Signal': result.get('macd_signal', ''),
                '50-Day MA': result.get('ma_50', 0),
                '200-Day MA': result.get('ma_200', 0),
                'Bollinger Position': result.get('bollinger_position', ''),
                'Support Level': result.get('support', 0),
                'Resistance Level': result.get('resistance', 0)
            })

    detailed_df = pd.DataFrame(detailed_data)
    sort_column = 'Consensus Score' if is_consensus else 'Overall Score'
    if sort_column in detailed_df.columns:
        detailed_df = detailed_df.sort_values(sort_column, ascending=False)
    detailed_df.to_excel(writer, sheet_name='Detailed_Analysis', index=False)

def create_technical_sheet(results, writer):
    """Create technical indicators sheet"""
    if not results:
        empty_df = pd.DataFrame({'Message': ['No technical data available']})
        empty_df.to_excel(writer, sheet_name='Technical_Indicators', index=False)
        return

    technical_data = []
    is_consensus = 'technical' in results[0]

    for result in results:
        if is_consensus:
            tech = result.get('technical', {})
            mom = result.get('momentum', {})
            current_price = result.get('current_price', 0)
            ma50 = mom.get('ma_50') or result.get('ma_50') or 0
            ma200 = mom.get('ma_200') or result.get('ma_200') or 0

            def pct_diff(price, ma):
                if not price or not ma:
                    return None
                return ((price / ma) - 1) * 100

            technical_data.append({
                'Symbol': result.get('symbol', ''),
                'RSI': mom.get('rsi'),
                'Price Trend': mom.get('price_trend'),
                'MACD': tech.get('macd'),
                'MACD Signal': tech.get('macd_signal'),
                'MACD Histogram': tech.get('macd_hist'),
                'Bollinger Position %': tech.get('bollinger_position'),
                'Bollinger Upper': tech.get('bollinger_upper'),
                'Bollinger Lower': tech.get('bollinger_lower'),
                'Support': tech.get('support'),
                'Resistance': tech.get('resistance'),
                'Volume SMA': tech.get('volume_sma'),
                'Volume Trend': mom.get('volume_trend'),
                'Price vs MA50 %': pct_diff(current_price, ma50),
                'Price vs MA200 %': pct_diff(current_price, ma200),
                'Momentum Score': result.get('momentum_score', mom.get('score')),
                'Technical Score': tech.get('score'),
                'Technical Grade': tech.get('grade', 'N/A')
            })
        else:
            signals = result.get('signals', {})
            current_price = result.get('current_price', 0)
            ma50 = result.get('ma_50', 0)
            ma200 = result.get('ma_200', 0)

            technical_data.append({
                'Symbol': result.get('symbol', ''),
                'RSI': result.get('rsi', 0),
                'RSI Signal': signals.get('rsi_signal', ''),
                'MACD': result.get('macd', 0),
                'MACD Signal': signals.get('macd_signal', ''),
                'Stochastic %K': result.get('stoch_k', 0),
                'Stochastic %D': result.get('stoch_d', 0),
                'Williams %R': result.get('williams_r', 0),
                'CCI': result.get('cci', 0),
                'ADX': result.get('adx', 0),
                'Bollinger Upper': result.get('bb_upper', 0),
                'Bollinger Lower': result.get('bb_lower', 0),
                'Bollinger Position': result.get('bollinger_position', ''),
                'Volume SMA': result.get('volume_sma', 0),
                'Price vs MA50': f"{((current_price / ma50 - 1) * 100):.1f}%" if ma50 else None,
                'Price vs MA200': f"{((current_price / ma200 - 1) * 100):.1f}%" if ma200 else None,
                'Support': result.get('support', 0),
                'Resistance': result.get('resistance', 0),
                'Technical Score': result.get('technical_score', 0)
            })

    technical_df = pd.DataFrame(technical_data)
    sort_col = 'Technical Score' if 'Technical Score' in technical_df.columns else None
    if sort_col:
        technical_df = technical_df.sort_values(sort_col, ascending=False)
    technical_df.to_excel(writer, sheet_name='Technical_Indicators', index=False)

def create_risk_analysis_sheet(results, writer):
    """Create risk analysis sheet"""
    if not results:
        empty_df = pd.DataFrame({'Message': ['No risk data available']})
        empty_df.to_excel(writer, sheet_name='Risk_Analysis', index=False)
        return

    risk_data = []
    is_consensus = 'risk' in results[0]

    for result in results:
        if is_consensus:
            risk = result.get('risk', {})
            fund = result.get('fundamentals', {})
            risk_data.append({
                'Symbol': result.get('symbol', ''),
                'Risk Level': risk.get('risk_level', result.get('risk_level', 'Unknown')),
                'Risk Score': risk.get('score', 0),
                'Risk Grade': risk.get('grade', 'N/A'),
                'Beta': risk.get('beta'),
                'Volatility %': risk.get('volatility'),
                'Sharpe Ratio': risk.get('sharpe_ratio'),
                'Max Drawdown %': risk.get('max_drawdown'),
                'VaR 95% %': risk.get('var_95'),
                'Debt/Equity': fund.get('debt_equity'),
                'Recommendation': result.get('recommendation', ''),
                'Consensus Score': result.get('consensus_score', result.get('overall_score', 0))
            })
        else:
            risk_data.append({
                'Symbol': result.get('symbol', ''),
                'Risk Level': result.get('risk_level', ''),
                'Beta': result.get('beta', 0),
                'Volatility': result.get('volatility', 0),
                'Sharpe Ratio': result.get('sharpe_ratio', 0),
                'Max Drawdown': result.get('max_drawdown', 0),
                'VaR (95%)': result.get('var_95', 0),
                'Debt/Equity': result.get('debt_equity', 0),
                'Current Ratio': result.get('current_ratio', 0),
                'Quick Ratio': result.get('quick_ratio', 0),
                'Interest Coverage': result.get('interest_coverage', 0),
                'Recommendation': result.get('recommendation', ''),
                'Overall Score': result.get('overall_score', 0)
            })

    risk_df = pd.DataFrame(risk_data)
    sort_col = 'Risk Score' if 'Risk Score' in risk_df.columns else None
    if sort_col:
        risk_df = risk_df.sort_values(sort_col, ascending=False)
    risk_df.to_excel(writer, sheet_name='Risk_Analysis', index=False)

def create_sector_analysis_sheet(results, writer):
    """Create sector analysis sheet"""
    if not results:
        empty_df = pd.DataFrame({'Message': ['No sector data available']})
        empty_df.to_excel(writer, sheet_name='Sector_Analysis', index=False)
        return

    sector_summary = {}
    consensus_format = 'consensus_score' in results[0] or 'overall_score' in results[0]

    for result in results:
        sector = result.get('sector', 'Unknown')
        sector_stats = sector_summary.setdefault(sector, {
            'count': 0,
            'strong_buy': 0,
            'buy': 0,
            'score_total': 0,
            'stocks': []
        })

        sector_stats['count'] += 1
        score_val = result.get('consensus_score') if consensus_format else result.get('overall_score', 0)
        sector_stats['score_total'] += score_val if score_val is not None else 0
        sector_stats['stocks'].append(result.get('symbol', ''))

        recommendation = result.get('recommendation', '')
        if recommendation == 'STRONG BUY':
            sector_stats['strong_buy'] += 1
        elif recommendation in ('BUY', 'WEAK BUY'):
            sector_stats['buy'] += 1

    sector_data = []
    for sector, data in sector_summary.items():
        count = data['count'] or 1
        sector_data.append({
            'Sector': sector,
            'Total Stocks': data['count'],
            'Strong Buy': data['strong_buy'],
            'Buy / Weak Buy': data['buy'],
            'Buy Rate %': f"{(data['strong_buy'] + data['buy']) / count * 100:.1f}%",
            'Avg Score': data['score_total'] / count,
            'Top Stocks': ', '.join(data['stocks'][:5])
        })

    sector_df = pd.DataFrame(sector_data)
    
    # Sort by number of stocks (using 'Total Stocks' from the DataFrame)
    if 'Total Stocks' in sector_df.columns:
        sector_df = sector_df.sort_values('Total Stocks', ascending=False)
    
    # Filter out empty/unknown sectors if they clutter the report, unless they have stocks
    sector_df = sector_df[~((sector_df['Sector'].isin(['Unknown', 'N/A'])) & (sector_df['Total Stocks'] == 0))]
    
    # If practically no sector data exists, add an explanatory row
    if sector_df.empty or (len(sector_df) == 1 and sector_df.iloc[0]['Sector'] == 'Unknown' and sector_df.iloc[0]['Total Stocks'] > 0):
         # If there's only 'Unknown' but it has stocks, keep it. Otherwise, show message.
         if sector_df.empty:
             sector_df = pd.DataFrame({'Message': ['Sector data unavailable from current data sources.']})
         else: # Only 'Unknown' sector with stocks, clarify it's uncategorized
             sector_df.loc[sector_df['Sector'] == 'Unknown', 'Sector'] = 'Uncategorized'
             sector_df = sector_df.rename(columns={'Sector': 'Sector (Uncategorized)'})
             
    sector_df.to_excel(writer, sheet_name='Sector_Analysis', index=False)

def create_performance_sheet(results, writer):
    """Create performance metrics sheet"""
    if not results:
        empty_df = pd.DataFrame({'Metric': ['No data available'], 'Value': ['']})
        empty_df.to_excel(writer, sheet_name='Performance_Metrics', index=False)
        return

    consensus_mode = 'consensus_score' in results[0] or 'quality_score' in results[0]

    if consensus_mode:
        scores = [r.get('consensus_score', 0) for r in results]
        quality_scores = [r.get('quality_score', 0) for r in results]
        confidences = [r.get('confidence', 0) * 100 for r in results]

        performance_data = {
            'Metric': [
                'Total Consensus Picks',
                'Average Consensus Score',
                'Median Consensus Score',
                'Consensus Score Std Dev',
                'Top 10% Consensus Threshold',
                'Bottom 10% Consensus Threshold',
                'Average Quality Score',
                'Median Quality Score',
                'Average Confidence %',
                '5/5 Agreement Count',
                '4/5 Agreement Count',
                '3/5 Agreement Count',
                '2/5 Agreement Count',
                'Ultimate Buy Count',
                'Strong Buy Count',
                'Buy Count',
                'Weak Buy Count',
                'Low Risk Count',
                'Medium Risk Count',
                'High Risk Count'
            ],
            'Value': [
                len(results),
                f"{np.mean(scores):.2f}",
                f"{np.median(scores):.2f}",
                f"{np.std(scores):.2f}",
                f"{np.percentile(scores, 90):.2f}",
                f"{np.percentile(scores, 10):.2f}",
                f"{np.mean(quality_scores):.2f}",
                f"{np.median(quality_scores):.2f}",
                f"{np.mean(confidences):.1f}%",
                len([r for r in results if r.get('strategies_agreeing') == 5]),
                len([r for r in results if r.get('strategies_agreeing') == 4]),
                len([r for r in results if r.get('strategies_agreeing') == 3]),
                len([r for r in results if r.get('strategies_agreeing') == 2]),
                len([r for r in results if r.get('recommendation') == 'ULTIMATE BUY']),
                len([r for r in results if r.get('recommendation') == 'STRONG BUY']),
                len([r for r in results if r.get('recommendation') == 'BUY']),
                len([r for r in results if r.get('recommendation') == 'WEAK BUY']),
                len([r for r in results if r.get('risk_level') == 'Low']),
                len([r for r in results if r.get('risk_level') == 'Medium']),
                len([r for r in results if r.get('risk_level') == 'High'])
            ]
        }
    else:
        scores = [r.get('overall_score', 0) for r in results]
        predictions = [r.get('prediction', 0) for r in results]

        performance_data = {
            'Metric': [
                'Total Stocks Analyzed',
                'Average Overall Score',
                'Median Overall Score',
                'Standard Deviation',
                'Top 10% Threshold',
                'Bottom 10% Threshold',
                'Average Predicted Return',
                'Max Predicted Return',
                'Min Predicted Return',
                'Strong Buy Count',
                'Buy Count',
                'Hold Count',
                'Sell Count',
                'High Risk Count',
                'Medium Risk Count',
                'Low Risk Count'
            ],
            'Value': [
                len(results),
                f"{np.mean(scores):.2f}",
                f"{np.median(scores):.2f}",
                f"{np.std(scores):.2f}",
                f"{np.percentile(scores, 90):.2f}",
                f"{np.percentile(scores, 10):.2f}",
                f"{np.mean(predictions) * 100:.2f}%",
                f"{np.max(predictions) * 100:.2f}%",
                f"{np.min(predictions) * 100:.2f}%",
                len([r for r in results if r.get('recommendation') == 'STRONG BUY']),
                len([r for r in results if r.get('recommendation') == 'BUY']),
                len([r for r in results if r.get('recommendation') == 'HOLD']),
                len([r for r in results if 'SELL' in r.get('recommendation', '')]),
                len([r for r in results if r.get('risk_level') == 'High']),
                len([r for r in results if r.get('risk_level') == 'Medium']),
                len([r for r in results if r.get('risk_level') == 'Low'])
            ]
        }

    performance_df = pd.DataFrame(performance_data)
    performance_df.to_excel(writer, sheet_name='Performance_Metrics', index=False)

def add_excel_export_to_streamlit(results):
    """Add Excel export functionality to Streamlit app"""
    
    import streamlit as st
    
    if results and len(results) > 0:
        st.markdown("---")
        st.subheader("📊 Export Results")
        
        col1, col2, col3 = st.columns([2, 1, 1])
        
        with col1:
            st.markdown("**Export your analysis results to Excel for further analysis and portfolio management.**")
        
        with col2:
            export_filename = st.text_input("Filename (optional)", placeholder="SmartTrade_Analysis")
        
        with col3:
            if st.button("📥 Export to Excel", type="primary"):
                # Add timestamp if no filename provided
                if not export_filename:
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    export_filename = f"SmartTrade_Analysis_{timestamp}"
                
                # Ensure .xlsx extension
                if not export_filename.endswith('.xlsx'):
                    export_filename += '.xlsx'
                
                # Export to Excel
                filename, message = export_analysis_to_excel(results, filename=export_filename)
                
                if filename:
                    st.success(f"✅ {message}")
                    
                    # Provide download link
                    if os.path.exists(filename):
                        with open(filename, 'rb') as f:
                            st.download_button(
                                label="📥 Download Excel File",
                                data=f.read(),
                                file_name=filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                else:
                    st.error(f"❌ {message}")
        
        # Show export details
        with st.expander("📋 Export Details"):
            st.markdown("""
            **Excel file includes 8 sheets:**
            - **Summary**: Overview and key metrics
            - **Strong Buy**: Top recommendations only
            - **All Buy Signals**: All buy recommendations
            - **Detailed Analysis**: Complete analysis data
            - **Technical Indicators**: All technical metrics
            - **Risk Analysis**: Risk assessment data
            - **Sector Analysis**: Sector breakdown
            - **Performance Metrics**: Statistical analysis
            """)

if __name__ == "__main__":
    # Test the export functionality
    sample_results = [
        {
            'symbol': 'AAPL',
            'company_name': 'Apple Inc.',
            'recommendation': 'STRONG BUY',
            'current_price': 175.50,
            'prediction': 0.15,
            'confidence': 0.85,
            'overall_score': 85,
            'technical_score': 80,
            'fundamental_score': 90,
            'risk_level': 'Medium',
            'sector': 'Technology'
        }
    ]
    
    filename, message = export_analysis_to_excel(sample_results)
    print(f"Test export: {message}")
